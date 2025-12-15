"""
Excel 工具模块 - 文件读取和Sheet处理
"""
import pandas as pd
from typing import List, Optional
import os


def get_sheet_names(filepath: str) -> List[str]:
    """
    获取Excel文件中的所有Sheet名称
    
    Args:
        filepath: Excel文件路径
    
    Returns:
        Sheet名称列表
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    # 使用openpyxl引擎（支持xlsx）或xlrd引擎（支持xls）
    if ext == ".xls":
        # 旧版xls格式使用xlrd
        try:
            import xlrd
            workbook = xlrd.open_workbook(filepath)
            return workbook.sheet_names()
        except ImportError:
            # 如果没有xlrd，尝试用pandas
            excel_file = pd.ExcelFile(filepath, engine='xlrd')
            return excel_file.sheet_names
    else:
        # xlsx/xlsm格式使用openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets


def load_excel(filepath: str, sheet_name: str, 
               header_row: int = 0,
               skip_rows: Optional[int] = None) -> pd.DataFrame:
    """
    加载Excel数据为DataFrame
    
    Args:
        filepath: Excel文件路径
        sheet_name: Sheet名称
        header_row: 表头行索引（0开始）
        skip_rows: 跳过行数
    
    Returns:
        DataFrame
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    read_kwargs = {
        "sheet_name": sheet_name,
        "header": header_row,
    }
    
    if skip_rows:
        read_kwargs["skiprows"] = skip_rows
    
    # 根据扩展名选择引擎
    if ext == ".xls":
        try:
            read_kwargs["engine"] = "xlrd"
        except:
            pass
    else:
        read_kwargs["engine"] = "openpyxl"
    
    df = pd.read_excel(filepath, **read_kwargs)
    
    # 清理数据
    df = clean_dataframe(df)
    
    return df


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    清理DataFrame
    
    - 去除列名首尾空格
    - 去除全空行
    - 字符串列去除首尾空格
    """
    # 清理列名
    df.columns = [str(col).strip() if col is not None else f"Column_{i}" 
                  for i, col in enumerate(df.columns)]
    
    # 去除全空行
    df = df.dropna(how="all")
    
    # 字符串列去空格
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].astype(str).str.strip()
        # 将 'nan' 字符串转回 NaN
        df[col] = df[col].replace("nan", pd.NA)
    
    return df


def detect_header_row(filepath: str, sheet_name: str, max_scan: int = 10) -> int:
    """
    自动检测表头行
    
    扫描前N行，找到非空列最多的行作为表头
    
    Args:
        filepath: Excel文件路径
        sheet_name: Sheet名称
        max_scan: 最大扫描行数
    
    Returns:
        表头行索引
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    
    # 读取前N行（无表头）
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, 
                          header=None, nrows=max_scan, engine=engine)
    except:
        df = pd.read_excel(filepath, sheet_name=sheet_name, 
                          header=None, nrows=max_scan)
    
    best_row = 0
    max_non_empty = 0
    
    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        non_empty = row.notna().sum()
        
        # 如果某行有很多非空值且看起来像文本（表头）
        if non_empty > max_non_empty:
            # 检查是否是有效表头（大部分是字符串）
            str_count = sum(1 for v in row if isinstance(v, str) and v.strip())
            if str_count >= non_empty * 0.5:
                max_non_empty = non_empty
                best_row = i
    
    return best_row


def preview_data(filepath: str, sheet_name: str, 
                 rows: int = 5, header_row: int = 0) -> dict:
    """
    预览数据
    
    Args:
        filepath: Excel文件路径
        sheet_name: Sheet名称
        rows: 预览行数
        header_row: 表头行
    
    Returns:
        {
            "columns": [...],
            "data": [{...}, {...}],
            "total_rows": 1000
        }
    """
    df = load_excel(filepath, sheet_name, header_row=header_row)
    
    return {
        "columns": list(df.columns),
        "data": df.head(rows).to_dict('records'),
        "total_rows": len(df)
    }

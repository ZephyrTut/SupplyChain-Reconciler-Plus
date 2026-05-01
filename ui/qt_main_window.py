"""
PyQt6 主窗口 - 供应链对账系统
使用 qt-material 主题
"""
import os
import sys
from typing import Optional
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QStackedWidget,
    QPushButton, QLabel, QComboBox, QFrame, QFileDialog, QMessageBox,
    QSplitter, QSizePolicy, QApplication
)
from PyQt6.QtCore import Qt, QMimeData, pyqtSignal
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QFont, QIcon, QWheelEvent
import pandas as pd

from config.settings import APP_NAME, APP_VERSION
from utils.excel_utils import get_sheet_names, load_excel, extract_unique_values
from utils.storage import load_templates, save_template, delete_template
from core.compare_engine import CompareEngine
from core.export_engine import ExportEngine


class NoScrollComboBox(QComboBox):
    """禁用鼠标滚轮的下拉框，避免干扰外部滚动"""
    
    def wheelEvent(self, event: QWheelEvent):
        event.ignore()


class FileDropCard(QFrame):
    """文件拖拽卡片组件"""
    
    file_dropped = pyqtSignal(str)  # 文件路径信号
    sheet_changed = pyqtSignal(str)  # Sheet变更信号
    
    def __init__(self, title: str, description: str, compact: bool = False, parent=None):
        super().__init__(parent)
        self.compact = compact
        self.filepath = ""
        self.setAcceptDrops(True)
        self.setObjectName("fileDropCard")
        self._setup_ui(title, description)
        
    def _setup_ui(self, title: str, description: str):
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(6 if self.compact else 10)
        
        # 图标（紧凑模式下缩小）
        icon_size = 24 if self.compact else 32
        icon_label = QLabel("📁")
        icon_label.setFont(QFont("Segoe UI Emoji", icon_size))
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)
        
        # 标题
        title_size = 11 if self.compact else 12
        title_label = QLabel(title)
        title_label.setFont(QFont("Microsoft YaHei", title_size, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 描述
        desc_label = QLabel(description)
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_label.setWordWrap(True)
        desc_style = "color: #666; font-size: 11px;" if self.compact else "color: #666;"
        desc_label.setStyleSheet(desc_style)
        layout.addWidget(desc_label)
        
        # 文件名显示
        self.file_label = QLabel("未选择文件")
        self.file_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.file_label.setStyleSheet("color: #999; font-style: italic;")
        layout.addWidget(self.file_label)
        
        # Sheet选择下拉框（初始隐藏）
        self.sheet_combo = NoScrollComboBox()
        self.sheet_combo.setVisible(False)
        self.sheet_combo.setMinimumWidth(150)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        layout.addWidget(self.sheet_combo, alignment=Qt.AlignmentFlag.AlignCenter)
        
        # 选择按钮
        self.select_btn = QPushButton("选择文件")
        self.select_btn.setObjectName("selectFileBtn")
        layout.addWidget(self.select_btn)
        
        # 样式（紧凑模式下减小最小高度和内边距）
        min_height = 150 if self.compact else 200
        padding = 12 if self.compact else 20
        self.setStyleSheet(f"""
            #fileDropCard {{
                border: 2px dashed #ccc;
                border-radius: 10px;
                background-color: #fafafa;
                min-height: {min_height}px;
                padding: {padding}px;
            }}
            #fileDropCard:hover {{
                border-color: #2196F3;
                background-color: #f0f7ff;
            }}
        """)
        self._default_style = self.styleSheet()
        
    def _on_sheet_changed(self, sheet_name: str):
        """Sheet选择变更"""
        if sheet_name and self.filepath:
            self.sheet_changed.emit(sheet_name)
        
    def set_file(self, filepath: str, sheets: list = None):
        """设置文件路径和可用Sheet"""
        self.filepath = filepath
        filename = os.path.basename(filepath)
        self.file_label.setText(f"✓ {filename}")
        self.file_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        
        # 更新Sheet下拉框
        if sheets and len(sheets) > 1:
            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            self.sheet_combo.blockSignals(False)
            self.sheet_combo.setVisible(True)
        else:
            self.sheet_combo.setVisible(False)
    
    def get_selected_sheet(self) -> str:
        """获取当前选中的Sheet"""
        if self.sheet_combo.isVisible():
            return self.sheet_combo.currentText()
        return ""
        
    def clear(self):
        """清空文件"""
        self.filepath = ""
        self.file_label.setText("未选择文件")
        self.file_label.setStyleSheet("color: #999; font-style: italic;")
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().lower().endswith(('.xlsx', '.xls', '.xlsm')):
                event.acceptProposedAction()
                min_height = 150 if self.compact else 200
                padding = 12 if self.compact else 20
                self.setStyleSheet(f"""
                    #fileDropCard {{
                        border: 2px dashed #2196F3;
                        border-radius: 10px;
                        background-color: #e3f2fd;
                        min-height: {min_height}px;
                        padding: {padding}px;
                    }}
                """)
                
    def dragLeaveEvent(self, event):
        self.setStyleSheet(self._default_style)
        
    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            filepath = urls[0].toLocalFile()
            if filepath.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                self.file_dropped.emit(filepath)
        self.dragLeaveEvent(event)


class StepIndicator(QWidget):
    """步骤指示器"""
    
    def __init__(self, compact: bool = False, parent=None):
        super().__init__(parent)
        self.current_step = 0
        self.compact = compact
        self._setup_ui()
        
    def _setup_ui(self):
        layout = QHBoxLayout(self)
        v_margin = 6 if self.compact else 10
        layout.setContentsMargins(20, v_margin, 20, v_margin)
        
        self.steps = []
        step_titles = ["导入文件", "配置规则", "执行对账"]
        
        for i, title in enumerate(step_titles):
            step_widget = self._create_step(i + 1, title)
            self.steps.append(step_widget)
            layout.addWidget(step_widget)
            
            # 添加连接线
            if i < len(step_titles) - 1:
                line = QFrame()
                line.setFrameShape(QFrame.Shape.HLine)
                line.setFixedHeight(2)
                line.setStyleSheet("background-color: #ddd;")
                layout.addWidget(line, 1)
                
    def _create_step(self, num: int, title: str) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(3 if self.compact else 5)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # 步骤圆圈（紧凑模式下缩小）
        size = 28 if self.compact else 36
        font_size = 11 if self.compact else 14
        circle = QLabel(str(num))
        circle.setFixedSize(size, size)
        circle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        circle.setFont(QFont("Arial", font_size, QFont.Weight.Bold))
        circle.setObjectName(f"stepCircle_{num}")
        layout.addWidget(circle, alignment=Qt.AlignmentFlag.AlignCenter)
        
        # 步骤标题
        label = QLabel(title)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setObjectName(f"stepLabel_{num}")
        if self.compact:
            label.setStyleSheet("font-size: 12px;")
        layout.addWidget(label)
        
        return widget
        
    def set_step(self, step: int):
        """设置当前步骤 (1-3)"""
        self.current_step = step
        radius = 14 if self.compact else 18
        for i in range(3):
            circle = self.findChild(QLabel, f"stepCircle_{i+1}")
            label = self.findChild(QLabel, f"stepLabel_{i+1}")
            
            if i + 1 < step:
                # 已完成
                circle.setStyleSheet(f"""
                    background-color: #4CAF50;
                    color: white;
                    border-radius: {radius}px;
                """)
                label.setStyleSheet("color: #4CAF50; font-weight: bold;")
            elif i + 1 == step:
                # 当前
                circle.setStyleSheet(f"""
                    background-color: #2196F3;
                    color: white;
                    border-radius: {radius}px;
                """)
                label.setStyleSheet("color: #2196F3; font-weight: bold;")
            else:
                # 未完成
                circle.setStyleSheet(f"""
                    background-color: #e0e0e0;
                    color: #666;
                    border-radius: {radius}px;
                """)
                label.setStyleSheet("color: #999;")


class QtMainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        
        # 数据
        self.manual_df: Optional[pd.DataFrame] = None
        self.system_df: Optional[pd.DataFrame] = None
        self.manual_path: str = ""
        self.system_path: str = ""
        self.result_df: Optional[pd.DataFrame] = None
        self.pivot_values: list = []  # 透视值列表
        self._template_reload_guard: bool = False  # 防止模板列表刷新时递归触发
        
        # 响应式尺寸计算
        self._calculate_responsive_sizes()
        
        self._setup_window()
        self._create_ui()
        self._connect_signals()
    
    def _calculate_responsive_sizes(self):
        """根据屏幕尺寸计算响应式参数"""
        screen = QApplication.primaryScreen().geometry()
        screen_width = screen.width()
        screen_height = screen.height()
        
        # 根据屏幕大小设置窗口尺寸（占屏幕80%，但有上下限）
        self.window_width = max(1000, min(1400, int(screen_width * 0.8)))
        self.window_height = max(650, min(900, int(screen_height * 0.85)))
        
        # 根据屏幕尺寸调整间距
        if screen_height < 800:
            self.spacing_scale = 0.6  # 小屏幕
        elif screen_height < 1000:
            self.spacing_scale = 0.8  # 中等屏幕
        else:
            self.spacing_scale = 1.0  # 大屏幕
            
        # 计算响应式间距
        self.content_margin = int(20 * self.spacing_scale)
        self.section_spacing = int(20 * self.spacing_scale)
        self.card_padding = int(15 * self.spacing_scale)
        
    def _setup_window(self):
        """设置窗口属性"""
        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION}")
        self.setMinimumSize(960, 600)
        self.resize(self.window_width, self.window_height)
        
        # 全局样式表 - 浅色主题，确保文字清晰可读
        self.setStyleSheet("""
            /* 主窗口背景 */
            QMainWindow {
                background-color: #f5f5f5;
            }
            
            /* 下拉框 - 浅色背景深色文字 */
            QComboBox {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 6px 10px;
                min-height: 26px;
                font-size: 13px;
            }
            QComboBox:hover {
                border-color: #2196F3;
            }
            QComboBox:focus {
                border-color: #2196F3;
                outline: none;
            }
            QComboBox::drop-down {
                border: none;
                width: 24px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #666;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                color: #333333;
                selection-background-color: #2196F3;
                selection-color: white;
                border: 1px solid #ccc;
                padding: 4px;
            }
            
            /* 输入框 - 浅色背景深色文字 */
            QLineEdit {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 6px 10px;
                min-height: 26px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #2196F3;
            }
            QLineEdit::placeholder {
                color: #999;
            }
            
            /* 滚动条样式 */
            QScrollBar:vertical {
                background-color: #f0f0f0;
                width: 12px;
                border-radius: 6px;
                margin: 2px;
            }
            QScrollBar::handle:vertical {
                background-color: #c0c0c0;
                border-radius: 5px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #a0a0a0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QScrollBar:horizontal {
                background-color: #f0f0f0;
                height: 12px;
                border-radius: 6px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal {
                background-color: #c0c0c0;
                border-radius: 5px;
                min-width: 30px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #a0a0a0;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
            
            /* 标签默认样式 */
            QLabel {
                color: #333333;
            }
            
            /* 消息框样式 - 确保文字可见 */
            QMessageBox {
                background-color: #ffffff;
            }
            QMessageBox QLabel {
                color: #333333;
                font-size: 13px;
            }
            QMessageBox QPushButton {
                background-color: #f5f5f5;
                color: #333333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 6px 16px;
                min-width: 70px;
            }
            QMessageBox QPushButton:hover {
                background-color: #e8e8e8;
            }
            QMessageBox QPushButton:default {
                background-color: #2196F3;
                color: white;
                border: none;
            }
            QMessageBox QPushButton:default:hover {
                background-color: #1976D2;
            }
            
            /* 分组框样式 */
            QGroupBox {
                color: #333333;
                font-weight: bold;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #333333;
            }
        """)
        
        # 居中显示
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)
        
    def _create_ui(self):
        """创建UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # 顶部栏
        self._create_header(main_layout)
        
        # 步骤指示器（响应式）
        self.step_indicator = StepIndicator(compact=self.spacing_scale < 1.0)
        main_layout.addWidget(self.step_indicator)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: #e0e0e0;")
        main_layout.addWidget(line)
        
        # 内容区域（堆叠窗口）
        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget, 1)
        
        # 创建三个步骤页面
        self._create_step1_page()
        self._create_step2_page()
        self._create_step3_page()
        
        # 底部按钮栏
        self._create_footer(main_layout)
        
        # 设置初始步骤
        self._show_step(1)
        
    def _create_header(self, parent_layout: QVBoxLayout):
        """创建顶部栏"""
        header = QFrame()
        header.setStyleSheet("background-color: #fff; border-bottom: 1px solid #e0e0e0;")
        header_layout = QHBoxLayout(header)
        v_margin = 6 if self.spacing_scale < 1.0 else 10
        header_layout.setContentsMargins(15, v_margin, 15, v_margin)
        
        # Logo和标题（紧凑模式缩小字体）
        title_size = 13 if self.spacing_scale < 1.0 else 16
        title = QLabel(f"📊 {APP_NAME}")
        title.setFont(QFont("Microsoft YaHei", title_size, QFont.Weight.Bold))
        title.setStyleSheet("color: #1976D2;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # 模板选择
        template_label = QLabel("配置模板:")
        template_label.setStyleSheet("color: #333;")
        header_layout.addWidget(template_label)
        
        self.template_combo = NoScrollComboBox()
        combo_width = 160 if self.spacing_scale < 1.0 else 200
        self.template_combo.setMinimumWidth(combo_width)
        self.template_combo.addItem("(选择模板)")
        self._load_templates()
        header_layout.addWidget(self.template_combo)
        
        # 模板操作按钮样式
        btn_style = """
            QPushButton {
                background-color: #f5f5f5;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #e8e8e8;
                border-color: #ccc;
            }
            QPushButton:pressed {
                background-color: #ddd;
            }
        """
        
        self.save_template_btn = QPushButton("💾 保存")
        self.save_template_btn.setToolTip("保存当前配置为模板")
        self.save_template_btn.setStyleSheet(btn_style)
        header_layout.addWidget(self.save_template_btn)
        
        self.delete_template_btn = QPushButton("🗑️ 删除")
        self.delete_template_btn.setToolTip("删除选中的模板")
        self.delete_template_btn.setStyleSheet(btn_style)
        header_layout.addWidget(self.delete_template_btn)
        
        parent_layout.addWidget(header)
        
    def _create_step1_page(self):
        """创建步骤1：文件导入"""
        page = QWidget()
        layout = QHBoxLayout(page)
        margin = int(30 * self.spacing_scale)
        layout.setContentsMargins(margin, margin, margin, margin)
        layout.setSpacing(int(30 * self.spacing_scale))
        
        # 手工表卡片
        self.manual_card = FileDropCard(
            "手工表",
            "拖拽Excel文件到这里，或点击选择\n支持 .xlsx / .xls / .xlsm",
            compact=self.spacing_scale < 1.0
        )
        layout.addWidget(self.manual_card)
        
        # 系统表卡片
        self.system_card = FileDropCard(
            "系统表",
            "拖拽Excel文件到这里，或点击选择\n支持 .xlsx / .xls / .xlsm",
            compact=self.spacing_scale < 1.0
        )
        layout.addWidget(self.system_card)
        
        self.stacked_widget.addWidget(page)
        
    def _create_step2_page(self):
        """创建步骤2：配置规则"""
        page = QWidget()
        layout = QHBoxLayout(page)
        margin = int(15 * self.spacing_scale)
        layout.setContentsMargins(margin, margin, margin, margin)
        layout.setSpacing(int(15 * self.spacing_scale))
        
        # 左侧配置面板（紧凑模式）
        from ui.qt_config_panel import QtConfigPanel
        self.config_panel = QtConfigPanel(compact=self.spacing_scale < 1.0)
        
        # 右侧预览面板（紧凑模式）
        from ui.qt_result_preview import QtResultPreview
        self.result_preview = QtResultPreview(compact=self.spacing_scale < 1.0)
        
        # 使用分割器，根据屏幕调整比例
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.config_panel)
        splitter.addWidget(self.result_preview)
        # 左侧配置面板尽量紧凑，右侧预览区域更大
        left_size = 380 if self.spacing_scale < 1.0 else 420
        right_size = 620 if self.spacing_scale < 1.0 else 780
        splitter.setSizes([left_size, right_size])
        
        layout.addWidget(splitter)
        self.stacked_widget.addWidget(page)
        
    def _create_step3_page(self):
        """创建步骤3：结果展示"""
        page = QWidget()
        layout = QVBoxLayout(page)
        margin = int(15 * self.spacing_scale)
        layout.setContentsMargins(margin, margin, margin, margin)
        layout.setSpacing(int(10 * self.spacing_scale))
        
        # 结果统计
        stats_frame = QFrame()
        stats_frame.setStyleSheet("background-color: #f5f5f5; border-radius: 8px; padding: 10px;")
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setSpacing(10)
        
        self.stats_total = self._create_stat_card("总计", "0", "#2196F3")
        self.stats_match = self._create_stat_card("一致", "0", "#4CAF50")
        self.stats_diff = self._create_stat_card("差异", "0", "#FF9800")
        self.stats_missing = self._create_stat_card("缺失", "0", "#F44336")
        
        stats_layout.addWidget(self.stats_total)
        stats_layout.addWidget(self.stats_match)
        stats_layout.addWidget(self.stats_diff)
        stats_layout.addWidget(self.stats_missing)
        
        layout.addWidget(stats_frame)
        
        # 结果表格
        from ui.qt_result_preview import QtResultTable
        self.result_table = QtResultTable()
        layout.addWidget(self.result_table, 1)
        
        self.stacked_widget.addWidget(page)
        
    def _create_stat_card(self, title: str, value: str, color: str) -> QWidget:
        """创建统计卡片"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border-radius: 8px;
                border-left: 4px solid {color};
                padding: 10px;
            }}
        """)
        layout = QVBoxLayout(card)
        layout.setSpacing(5)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("color: #666;")
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        value_label.setStyleSheet(f"color: {color};")
        value_label.setObjectName(f"stat_{title}")
        layout.addWidget(value_label)
        
        return card
        
    def _create_footer(self, parent_layout: QVBoxLayout):
        """创建底部按钮栏"""
        footer = QFrame()
        footer.setStyleSheet("background-color: #fff; border-top: 1px solid #e0e0e0;")
        footer_layout = QHBoxLayout(footer)
        v_margin = 10 if self.spacing_scale < 1.0 else 15
        footer_layout.setContentsMargins(15, v_margin, 15, v_margin)
        
        # 左侧提示
        self.status_label = QLabel("请导入手工表和系统表")
        self.status_label.setStyleSheet("color: #666;")
        footer_layout.addWidget(self.status_label)
        
        footer_layout.addStretch()
        
        # 按钮尺寸根据屏幕调整
        btn_padding = "8px 20px" if self.spacing_scale < 1.0 else "10px 30px"
        
        # 导航按钮
        self.prev_btn = QPushButton("← 上一步")
        self.prev_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #ffffff;
                color: #2196F3;
                border: 2px solid #2196F3;
                padding: {btn_padding};
                border-radius: 5px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #e3f2fd;
            }}
        """)
        self.prev_btn.setVisible(False)
        footer_layout.addWidget(self.prev_btn)
        
        self.next_btn = QPushButton("下一步 →")
        self.next_btn.setObjectName("primaryBtn")
        self.next_btn.setStyleSheet(f"""
            #primaryBtn {{
                background-color: #2196F3;
                color: white;
                border: none;
                padding: {btn_padding};
                border-radius: 5px;
                font-weight: bold;
            }}
            #primaryBtn:hover {{
                background-color: #1976D2;
            }}
            #primaryBtn:disabled {{
                background-color: #ccc;
            }}
        """)
        footer_layout.addWidget(self.next_btn)
        
        self.run_btn = QPushButton("🚀 执行对账")
        self.run_btn.setObjectName("runBtn")
        self.run_btn.setStyleSheet(f"""
            #runBtn {{
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: {btn_padding};
                border-radius: 5px;
                font-weight: bold;
            }}
            #runBtn:hover {{
                background-color: #388E3C;
            }}
        """)
        self.run_btn.setVisible(False)
        footer_layout.addWidget(self.run_btn)
        
        self.export_btn = QPushButton("📥 导出Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setStyleSheet(f"""
            #exportBtn {{
                background-color: #FF9800;
                color: white;
                border: none;
                padding: {btn_padding};
                border-radius: 5px;
                font-weight: bold;
            }}
            #exportBtn:hover {{
                background-color: #F57C00;
            }}
        """)
        self.export_btn.setVisible(False)
        footer_layout.addWidget(self.export_btn)
        
        parent_layout.addWidget(footer)
        
    def _connect_signals(self):
        """连接信号槽"""
        # 文件选择
        self.manual_card.select_btn.clicked.connect(lambda: self._select_file("manual"))
        self.system_card.select_btn.clicked.connect(lambda: self._select_file("system"))
        self.manual_card.file_dropped.connect(lambda p: self._load_file(p, "manual"))
        self.system_card.file_dropped.connect(lambda p: self._load_file(p, "system"))
        
        # Sheet选择变更
        self.manual_card.sheet_changed.connect(lambda s: self._on_sheet_changed("manual", s))
        self.system_card.sheet_changed.connect(lambda s: self._on_sheet_changed("system", s))
        
        # 导航按钮
        self.prev_btn.clicked.connect(self._go_prev)
        self.next_btn.clicked.connect(self._go_next)
        self.run_btn.clicked.connect(self._run_comparison)
        self.export_btn.clicked.connect(self._export_results)
        
        # 模板
        self.template_combo.currentIndexChanged.connect(self._on_template_selected)
        self.save_template_btn.clicked.connect(self._save_template)
        self.delete_template_btn.clicked.connect(self._delete_template)
        
        # 配置变更
        self.config_panel.config_changed.connect(self._on_config_changed)
        
        # 导出预处理预览
        self.config_panel.export_preview_requested.connect(self._export_manual_preview)
        self.config_panel.export_system_requested.connect(self._export_system_preview)
        
    def _show_step(self, step: int):
        """显示指定步骤"""
        self.current_step = step
        self.step_indicator.set_step(step)
        self.stacked_widget.setCurrentIndex(step - 1)
        
        # 更新按钮状态
        self.prev_btn.setVisible(step > 1)
        self.next_btn.setVisible(step == 1)  # 只在步骤1显示下一步
        self.run_btn.setVisible(step == 2)  # 步骤2显示执行对账和导出
        self.export_btn.setVisible(step >= 2)  # 步骤2和3都显示导出
        
        # 更新状态提示和数据显示
        if step == 1:
            self._update_step1_status()
        elif step == 2:
            self.status_label.setText("配置主键和数值列后，点击执行对账")
            # 步骤2：更新预览
            if self.manual_df is not None and self.system_df is not None:
                config = self.config_panel.get_config()
                self.result_preview.update_preview(
                    self.manual_df,
                    self.system_df,
                    config
                )
        elif step == 3:
            self.status_label.setText("对账完成！可导出Excel结果")
            # 步骤3：确保结果表格已更新
            if self.result_df is not None:
                config = self.config_panel.get_config()
                self.result_table.set_data(self.result_df, config)
            
    def _update_step1_status(self):
        """更新步骤1状态"""
        manual_ok = self.manual_df is not None
        system_ok = self.system_df is not None
        
        if manual_ok and system_ok:
            self.status_label.setText(f"✓ 已导入: 手工表 {len(self.manual_df)}行, 系统表 {len(self.system_df)}行")
            self.next_btn.setEnabled(True)
        elif manual_ok:
            self.status_label.setText("✓ 已导入手工表, 请导入系统表")
            self.next_btn.setEnabled(False)
        elif system_ok:
            self.status_label.setText("请导入手工表, ✓ 已导入系统表")
            self.next_btn.setEnabled(False)
        else:
            self.status_label.setText("请导入手工表和系统表")
            self.next_btn.setEnabled(False)
            
    def _select_file(self, file_type: str):
        """选择文件"""
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            f"选择{'手工表' if file_type == 'manual' else '系统表'}",
            "",
            "Excel文件 (*.xlsx *.xls *.xlsm)"
        )
        if filepath:
            self._load_file(filepath, file_type)
            
    def _load_file(self, filepath: str, file_type: str, sheet_name: str = None):
        """加载文件"""
        try:
            sheets = get_sheet_names(filepath)
            card = self.manual_card if file_type == "manual" else self.system_card
            
            # 如果没有指定sheet，使用第一个
            if sheet_name is None:
                sheet_name = sheets[0]
                
            df = load_excel(filepath, sheet_name)
            
            if file_type == "manual":
                self.manual_df = df
                self.manual_path = filepath
                card.set_file(filepath, sheets)
            else:
                self.system_df = df
                self.system_path = filepath
                card.set_file(filepath, sheets)
                
            self._update_step1_status()
            self._refresh_config_panel_metadata()
                
        except Exception as e:
            from ui.qt_dialogs import show_error
            show_error(self, "导入失败", f"无法读取文件:\n{str(e)}")
            
    def _on_sheet_changed(self, file_type: str, sheet_name: str):
        """Sheet选择变更处理"""
        if not sheet_name:
            return
        filepath = self.manual_path if file_type == "manual" else self.system_path
        if filepath:
            try:
                df = load_excel(filepath, sheet_name)
                if file_type == "manual":
                    self.manual_df = df
                else:
                    self.system_df = df
                self._update_step1_status()
                self._refresh_config_panel_metadata()
            except Exception as e:
                from ui.qt_dialogs import show_warning
                show_warning(self, "加载失败", f"无法加载工作表:\n{str(e)}")

    def _refresh_config_panel_metadata(self):
        """刷新配置面板列信息与筛选唯一值。"""
        manual_columns = list(self.manual_df.columns) if self.manual_df is not None else []
        system_columns = list(self.system_df.columns) if self.system_df is not None else []

        self.config_panel.set_columns(manual_columns, system_columns)

        manual_unique_values = extract_unique_values(self.manual_df) if self.manual_df is not None else {}
        system_unique_values = extract_unique_values(self.system_df) if self.system_df is not None else {}

        self.config_panel.set_manual_unique_values(manual_unique_values)
        self.config_panel.set_system_unique_values(system_unique_values)
        self._update_auto_map_stats()

    def _build_filter_tuples(self, filters: list) -> list:
        """将配置中的筛选规则转换为引擎可用结构（兼容旧tuple格式）。"""
        items = []
        for f in filters or []:
            column = f.get("column")
            operator = f.get("operator")
            value = f.get("value")
            if column and operator:
                item = {"column": column, "operator": operator, "value": value}
                if f.get("target_filter"):
                    item["target_filter"] = f.get("target_filter")
                items.append(item)
        return items

    def _format_filter_item(self, item: dict) -> str:
        """格式化筛选规则用于导出显示。"""
        col = item.get("column", "")
        op = item.get("operator", "")
        val = item.get("value", "")
        text = f"{col} {op} '{val}'"

        target = item.get("target_filter")
        if isinstance(target, dict):
            t_col = target.get("column", "")
            t_op = target.get("operator", "")
            t_val = target.get("value", "")
            if t_col and t_op:
                text += f" (仅覆盖: {t_col} {t_op} '{t_val}')"

        return text
            
    def _go_prev(self):
        """上一步"""
        if self.current_step > 1:
            self._show_step(self.current_step - 1)
            
    def _go_next(self):
        """下一步"""
        if self.current_step < 3:
            self._show_step(self.current_step + 1)
            
    def _run_comparison(self):
        """执行对账"""
        loading = None
        try:
            config = self.config_panel.get_config()
            
            # 验证配置
            from ui.qt_dialogs import show_warning
            if not config.get("key_mappings"):
                show_warning(self, "配置不完整", "请至少配置一个主键映射")
                return
            if not config.get("value_mapping", {}).get("manual"):
                show_warning(self, "配置不完整", "请配置手工表数值列")
                return
                
            # 执行对账
            from ui.qt_dialogs import LoadingDialog
            loading = LoadingDialog("正在执行对账...", self)
            loading.show()
            QApplication.processEvents()
            
            # 准备配置参数
            key_mappings = config.get("key_mappings", [])
            manual_key_cols = [
                (k.get("manual") or k.get("manual_col"))
                for k in key_mappings
                if isinstance(k, dict) and (k.get("manual") or k.get("manual_col"))
            ]
            system_key_cols = [
                (k.get("system") or k.get("system_col"))
                for k in key_mappings
                if isinstance(k, dict) and (k.get("system") or k.get("system_col"))
            ]
            
            value_mapping = config.get("value_mapping", {})
            manual_val_col = value_mapping.get("manual", "") or value_mapping.get("manual_col", "")
            system_val_col = value_mapping.get("system", "") or value_mapping.get("system_col", "")
            
            # 透视列配置
            pivot_config = config.get("pivot_column", {})
            pivot_col = pivot_config.get("system", "") if isinstance(pivot_config, dict) else pivot_config
            
            # 列清洗规则
            clean_rules = config.get("clean_rules", [])
            
            # 手工表透视配置
            manual_pivot = config.get("manual_pivot", {})
            
            # 复制数据进行处理
            manual_data = self.manual_df.copy()
            system_data = self.system_df.copy()
            
            # 应用列清洗（仅手工表）
            if clean_rules:
                manual_data = CompareEngine.clean_column(manual_data, clean_rules)
            
            # 自动映射系统表零件号（仅影响副本）
            auto_map = config.get("system_auto_map", {})
            if auto_map.get("enabled"):
                system_data, _ = CompareEngine.auto_map_system_parts(
                    system_data,
                    manual_data,
                    auto_map
                )

            # 生成主键
            manual_with_key = CompareEngine.make_key(manual_data, manual_key_cols)
            system_with_key = CompareEngine.make_key(system_data, system_key_cols)
            
            # 准备筛选条件
            manual_filters = self._build_filter_tuples(config.get("manual_filters", []))
            manual_filter_exceptions = self._build_filter_tuples(config.get("manual_filter_exceptions", []))

            system_filters = self._build_filter_tuples(config.get("system_filters", []))
            system_filter_exceptions = self._build_filter_tuples(config.get("system_filter_exceptions", []))
            
            # 聚合数据
            # 手工表聚合 - 检查是否有手工表透视配置
            if manual_pivot and manual_pivot.get("pivot_column"):
                # 使用手工表透视聚合（区分出库/入库）
                manual_agg, out_cols, in_cols = CompareEngine.aggregate_manual_with_pivot(
                    manual_with_key, "__KEY__", manual_val_col,
                    manual_pivot,
                    filters=manual_filters,
                    filter_exceptions=manual_filter_exceptions
                )
                # 保存手工表透视信息（用于结果显示）
                self.manual_pivot_info = {"out_cols": out_cols, "in_cols": in_cols}
            else:
                # 普通聚合
                manual_agg, _ = CompareEngine.aggregate_data(
                    manual_with_key, "__KEY__", [manual_val_col] if manual_val_col else [],
                    filters=manual_filters,
                    filter_exceptions=manual_filter_exceptions
                )
                self.manual_pivot_info = None
            
            system_agg, pivot_values = CompareEngine.aggregate_data(
                system_with_key, "__KEY__", [system_val_col] if system_val_col else [],
                pivot_col=pivot_col if pivot_col else None,
                filters=system_filters,
                filter_exceptions=system_filter_exceptions
            )
            
            # 保存透视值
            self.pivot_values = pivot_values
            
            # 确定数值列名（merge_and_compare会将它们重命名为标准名称）
            manual_val = manual_val_col if manual_val_col else ""
            system_val = system_val_col if system_val_col else ""
            
            # 获取字母公式并转换为列名公式
            letter_formula = config.get("difference_formula", "")
            
            
            # 构建字母到列名的映射
            # 新列顺序: A=__KEY__, [B,C,D...=透视列], 系统总计, 手工数量, 差值, 比对状态
            letter_to_column = {}
            
            # 系统总计和透视列、手工数量
            if pivot_col and pivot_values:
                # 有透视列时：B,C,D=透视列，然后系统总计，然后手工数量
                letter_index = ord('B')  # 从B开始
                for pv in sorted(pivot_values):
                    letter_to_column[chr(letter_index)] = pv
                    letter_index += 1
                letter_to_column[chr(letter_index)] = "系统总计"
                letter_index += 1
                letter_to_column[chr(letter_index)] = "手工数量"
            else:
                # 无透视列：B=系统总计, C=手工数量
                letter_to_column["B"] = "系统总计"
                letter_to_column["C"] = "手工数量"
            
            
            # 将字母公式转换为列名公式
            column_formula = letter_formula
            # 按字母逆序替换（避免B被BB等部分匹配）
            for letter in sorted(letter_to_column.keys(), key=lambda x: ord(x), reverse=True):
                column_name = letter_to_column[letter]
                column_formula = column_formula.replace(letter, column_name)
            
            
            # 合并比对
            self.result_df = CompareEngine.merge_and_compare(
                manual_agg, system_agg, "__KEY__",
                manual_val, system_val,
                diff_formula=column_formula,
                pivot_values=pivot_values
            )
            
            if loading:
                loading.close()
                loading = None
            
            # 更新统计
            self._update_stats()
            
            # 更新结果表格（传入配置以显示公式）
            self.result_table.set_data(self.result_df, config)
            
            # 进入步骤3
            self._show_step(3)
            
        except Exception as e:
            if loading:
                loading.close()
            from ui.qt_dialogs import show_error
            show_error(self, "对账失败", f"执行对账时出错:\n{str(e)}")
            
    def _update_stats(self):
        """更新统计信息"""
        if self.result_df is None:
            return
            
        total = len(self.result_df)
        # 使用 str.startswith 来匹配状态（因为状态是 "✓ 一致" 这样的完整字符串）
        status_col = self.result_df['比对状态'].astype(str)
        match = len(self.result_df[status_col.str.startswith('✓')])
        diff = len(self.result_df[status_col.str.startswith('↕')])
        missing = len(self.result_df[status_col.str.startswith('✗')])
        
        
        # 查找标签
        stat_total = self.findChild(QLabel, "stat_总计")
        stat_match = self.findChild(QLabel, "stat_一致")
        stat_diff = self.findChild(QLabel, "stat_差异")
        stat_missing = self.findChild(QLabel, "stat_缺失")
        
        
        if stat_total:
            stat_total.setText(str(total))
        if stat_match:
            stat_match.setText(str(match))
        if stat_diff:
            stat_diff.setText(str(diff))
        if stat_missing:
            stat_missing.setText(str(missing))
        
    def _export_results(self):
        """导出结果"""
        if self.result_df is None:
            from ui.qt_dialogs import show_warning
            show_warning(self, "无数据", "没有对账结果可导出")
            return
            
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "保存对账结果",
            f"对账结果_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx)"
        )
        
        if filepath:
            try:
                config = self.config_panel.get_config()
                pivot_values = config.get("pivot_values", [])
                ExportEngine.export_results(filepath, self.result_df, pivot_values, config)
                
                from ui.qt_dialogs import show_info, show_error
                show_info(self, "导出成功", f"结果已保存到:\n{filepath}")
                
                # 打开文件夹
                os.startfile(os.path.dirname(filepath))
                
            except Exception as e:
                show_error(self, "导出失败", f"导出文件时出错:\n{str(e)}")
                
    def _load_templates(self):
        """加载模板列表"""
        templates = load_templates()
        self.template_combo.clear()
        self.template_combo.addItem("(选择模板)")
        for t in templates:
            self.template_combo.addItem(t.get("name", "未命名"), t)
            
    def _on_template_selected(self, index: int):
        """模板选择事件"""
        if self._template_reload_guard:
            return
        if index <= 0:
            return
        template = self.template_combo.itemData(index)
        if template:
            config = template.get("config", {})
            self.config_panel.set_config(config)
            # 模板加载后主动刷新预览与公式快速选择
            self._on_config_changed()
            # 模板加载后自动回写更新（基于当前两表数据）
            self._auto_writeback_loaded_template(template)

    def _auto_writeback_loaded_template(self, template: dict):
        """加载模板后自动回写更新（仅在两表都已导入时执行）。"""
        if self.manual_df is None or self.system_df is None:
            return

        template_name = template.get("name", "") if isinstance(template, dict) else ""
        if not template_name:
            return

        old_config = template.get("config", {}) if isinstance(template, dict) else {}
        new_config = self.config_panel.get_config()

        if not self._is_template_writeback_allowed(new_config):
            return

        # 没变化则不写盘
        if new_config == old_config:
            return

        ok = save_template(template_name, new_config)
        if not ok:
            return

        # 刷新模板列表并保持当前选中，避免UI显示旧itemData
        self._template_reload_guard = True
        try:
            self._load_templates()
            idx = self.template_combo.findText(template_name)
            if idx >= 0:
                self.template_combo.setCurrentIndex(idx)
        finally:
            self._template_reload_guard = False

    def _is_template_writeback_allowed(self, config: dict) -> bool:
        """避免用不完整配置覆盖模板。"""
        if not isinstance(config, dict):
            return False

        key_mappings = config.get("key_mappings", [])
        if not key_mappings:
            return False

        value_mapping = config.get("value_mapping", {})
        if not value_mapping.get("manual") or not value_mapping.get("system"):
            return False

        return True
            
    def _save_template(self):
        """保存模板"""
        from ui.qt_dialogs import InputDialog, show_info, show_warning
        
        # 获取已有模板名称列表
        templates = load_templates()
        existing_names = [t.get("name", "") for t in templates if t.get("name")]
        
        dialog = InputDialog(
            "保存模板", 
            "请输入模板名称:",
            "", 
            options=existing_names,  # 提供已有模板选项
            parent=self
        )
        if dialog.exec():
            name = dialog.get_text()
            if name:
                config = self.config_panel.get_config()
                ok = save_template(name, config)
                if not ok:
                    show_warning(self, "保存失败", "模板保存失败，请检查模板名称或文件权限")
                    return
                self._load_templates()
                # 选中新模板
                index = self.template_combo.findText(name)
                if index >= 0:
                    self.template_combo.setCurrentIndex(index)
                show_info(self, "保存成功", f"模板 '{name}' 已保存")
                
    def _delete_template(self):
        """删除模板"""
        from ui.qt_dialogs import show_warning, show_info, show_confirm
        
        index = self.template_combo.currentIndex()
        if index <= 0:
            show_warning(self, "提示", "请先选择要删除的模板")
            return
            
        template = self.template_combo.itemData(index)
        name = template.get("name", "")
        
        if show_confirm(self, "确认删除", f"确定要删除模板 '{name}' 吗？\n此操作不可恢复。", danger=True):
            template_id = template.get("id") or template.get("name")
            success, msg = delete_template(template_id)
            if success:
                self._load_templates()
                show_info(self, "删除成功", f"模板 '{name}' 已删除")
            else:
                show_warning(self, "删除失败", msg)
                
    def _on_config_changed(self):
        """配置变更事件"""
        # 更新预览
        if self.manual_df is not None and self.system_df is not None:
            config = self.config_panel.get_config()
            self.result_preview.update_preview(
                self.manual_df,
                self.system_df,
                config
            )
            
            # 获取列字母映射，更新配置面板的公式快速选择
            column_letters = self.result_preview.get_column_letters()
            if column_letters:
                self.config_panel.update_formula_options(column_letters)
        self._update_auto_map_stats()

    def _update_auto_map_stats(self):
        if self.manual_df is None or self.system_df is None:
            self.config_panel.set_auto_map_stats("映射统计: 未导入完整数据")
            return

        config = self.config_panel.get_config()
        auto_map = config.get("system_auto_map", {})
        if not auto_map.get("enabled"):
            self.config_panel.set_auto_map_stats("映射统计: 未启用")
            return

        from core.compare_engine import CompareEngine
        _, stats = CompareEngine.auto_map_system_parts(
            self.system_df,
            self.manual_df,
            auto_map,
            output_column="__MAPPED_PART__"
        )
        text = (
            f"映射统计: 候选{stats.get('candidates', 0)}行, "
            f"匹配{stats.get('matched', 0)}行, "
            f"歧义{stats.get('ambiguous', 0)}行, "
            f"未匹配{stats.get('unmatched', 0)}行"
        )
        self.config_panel.set_auto_map_stats(text)
    
    def _export_manual_preview(self):
        """导出手工表预处理预览（显示清洗和透视计算过程）"""
        from ui.qt_dialogs import show_warning, show_info
        
        if self.manual_df is None:
            show_warning(self, "无数据", "请先导入手工表")
            return
        
        config = self.config_panel.get_config()
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import os
        
        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        import pandas as pd
        import numpy as np
        
        def safe_value(val):
            """转换值为Excel兼容格式"""
            if pd.isna(val) or val is pd.NA:
                return ""
            if isinstance(val, (np.integer, np.floating)):
                return float(val) if np.isfinite(val) else ""
            return val
        
        # === Sheet1: 原始数据 ===
        ws1 = wb.active
        ws1.title = "1-原始数据"
        df_original = self.manual_df.copy()
        
        ws1.cell(row=1, column=1, value="【手工表原始数据】").font = Font(bold=True, size=12, color="0000FF")
        ws1.cell(row=2, column=1, value=f"共 {len(df_original)} 行数据")
        
        start_row = 4
        for c_idx, col in enumerate(df_original.columns, 1):
            cell = ws1.cell(row=start_row, column=c_idx, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")
            cell.border = thin_border
        for r_idx, row in enumerate(df_original.itertuples(index=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=safe_value(value))
                cell.border = thin_border
        
        # === Sheet2: 清洗后数据 ===
        ws2 = wb.create_sheet("2-清洗后数据")
        clean_rules = config.get("clean_rules", [])
        df_cleaned = df_original.copy()
        
        ws2.cell(row=1, column=1, value="【清洗规则】").font = Font(bold=True, size=12, color="FF0000")
        
        if clean_rules:
            for i, rule in enumerate(clean_rules):
                mode_text = f"{rule['column']}: {rule['mode']} 正则'{rule.get('regexes', [])}'"
                if rule.get('replace'):
                    mode_text += f" => '{rule['replace']}'"
                ws2.cell(row=2+i, column=1, value=f"规则{i+1}: {mode_text}")
            
            df_cleaned = CompareEngine.clean_column(df_cleaned, clean_rules)
            start_row = 4 + len(clean_rules)
        else:
            ws2.cell(row=2, column=1, value="（无清洗规则）")
            start_row = 4
        
        for c_idx, col in enumerate(df_cleaned.columns, 1):
            cell = ws2.cell(row=start_row, column=c_idx, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
            cell.border = thin_border
        for r_idx, row in enumerate(df_cleaned.itertuples(index=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=safe_value(value))
                cell.border = thin_border
        
        # === Sheet3: 透视计算结果 ===
        ws3 = wb.create_sheet("3-透视计算结果")
        manual_pivot = config.get("manual_pivot", {})
        
        ws3.cell(row=1, column=1, value="【手工表透视配置】").font = Font(bold=True, size=12, color="0000FF")
        
        if manual_pivot and manual_pivot.get("pivot_column"):
            pivot_col = manual_pivot.get("pivot_column", "")
            out_values = manual_pivot.get("out_values", [])
            in_values = manual_pivot.get("in_values", [])
            
            ws3.cell(row=2, column=1, value=f"透视列: {pivot_col}")
            ws3.cell(row=3, column=1, value=f"📤 出库值: {', '.join(out_values) if out_values else '(无)'}")
            ws3.cell(row=4, column=1, value=f"📥 入库值: {', '.join(in_values) if in_values else '(无)'}")
            ws3.cell(row=5, column=1, value="计算公式: 手工数量 = Σ出库 - Σ入库").font = Font(bold=True, color="FF6600")
            
            # 执行透视计算
            key_mappings = config.get("key_mappings", [])
            manual_key_cols = [
                (k.get("manual") or k.get("manual_col"))
                for k in key_mappings
                if isinstance(k, dict) and (k.get("manual") or k.get("manual_col"))
            ]
            value_mapping = config.get("value_mapping", {})
            manual_val_col = value_mapping.get("manual", "") or value_mapping.get("manual_col", "")
            
            if manual_key_cols and manual_val_col:
                manual_with_key = CompareEngine.make_key(df_cleaned, manual_key_cols)
                manual_filters = self._build_filter_tuples(config.get("manual_filters", []))
                manual_filter_exceptions = self._build_filter_tuples(config.get("manual_filter_exceptions", []))
                
                try:
                    pivot_df, out_cols, in_cols = CompareEngine.aggregate_manual_with_pivot(
                        manual_with_key,
                        "__KEY__",
                        manual_val_col,
                        manual_pivot,
                        filters=manual_filters,
                        filter_exceptions=manual_filter_exceptions
                    )
                    
                    start_row = 7
                    for c_idx, col in enumerate(pivot_df.columns, 1):
                        cell = ws3.cell(row=start_row, column=c_idx, value=col)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        if col in out_values:
                            cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")  # 蓝色-出库
                        elif col in in_values:
                            cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")  # 绿色-入库
                        elif col == "手工数量":
                            cell.fill = PatternFill(start_color="FFF3E0", fill_type="solid")  # 橙色-结果
                    
                    for r_idx, row in enumerate(pivot_df.itertuples(index=False), start_row + 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                            cell.border = thin_border
                except Exception as e:
                    ws3.cell(row=7, column=1, value=f"透视计算出错: {str(e)}")
            else:
                ws3.cell(row=7, column=1, value="（请先配置主键和数值列）")
        else:
            ws3.cell(row=2, column=1, value="（未配置手工表透视）")
        
        # 保存文件
        from PyQt6.QtWidgets import QFileDialog
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"手工表预处理预览_{timestamp}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出预处理预览", default_name, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb.save(file_path)
                show_info(self, "导出成功", f"预处理预览已保存:\n{file_path}")
                os.startfile(file_path)  # 自动打开
            except Exception as e:
                show_warning(self, "导出失败", f"保存文件时出错: {str(e)}")

    def _export_system_preview(self):
        """导出系统表预处理预览（显示筛选和透视计算过程）"""
        from ui.qt_dialogs import show_warning, show_info
        
        if self.system_df is None:
            show_warning(self, "无数据", "请先导入系统表")
            return
        
        config = self.config_panel.get_config()
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import os
        
        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        import pandas as pd
        import numpy as np
        
        def safe_value(val):
            """转换值为Excel兼容格式"""
            if pd.isna(val) or val is pd.NA:
                return ""
            if isinstance(val, (np.integer, np.floating)):
                return float(val) if np.isfinite(val) else ""
            return val
        
        # === Sheet1: 原始数据 ===
        ws1 = wb.active
        ws1.title = "1-原始数据"
        df_original = self.system_df.copy()
        df_for_processing = df_original.copy()
        map_stats_text = "映射统计: 未启用"
        auto_map = config.get("system_auto_map", {})
        if auto_map.get("enabled"):
            if self.manual_df is None:
                map_stats_text = "映射统计: 未导入手工表"
            else:
                df_for_processing, stats = CompareEngine.auto_map_system_parts(
                    df_for_processing,
                    self.manual_df,
                    auto_map
                )
                map_stats_text = (
                    f"映射统计: 候选{stats.get('candidates', 0)}行, "
                    f"匹配{stats.get('matched', 0)}行, "
                    f"歧义{stats.get('ambiguous', 0)}行, "
                    f"未匹配{stats.get('unmatched', 0)}行"
                )
        
        ws1.cell(row=1, column=1, value="【系统表原始数据】").font = Font(bold=True, size=12, color="2E7D32")
        ws1.cell(row=2, column=1, value=f"共 {len(df_original)} 行数据")
        
        start_row = 4
        for c_idx, col in enumerate(df_original.columns, 1):
            cell = ws1.cell(row=start_row, column=c_idx, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
            cell.border = thin_border
        for r_idx, row in enumerate(df_original.itertuples(index=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=safe_value(value))
                cell.border = thin_border
        
        # === Sheet2: 筛选后数据 ===
        ws2 = wb.create_sheet("2-筛选后数据")
        system_filters = config.get("system_filters", [])
        system_filter_exceptions = config.get("system_filter_exceptions", [])
        df_filtered = df_for_processing.copy()
        
        ws2.cell(row=1, column=1, value="【筛选规则】").font = Font(bold=True, size=12, color="FF0000")
        
        ws2.cell(row=2, column=1, value=map_stats_text)

        if system_filters or system_filter_exceptions:
            filter_tuples = self._build_filter_tuples(system_filters)
            exception_tuples = self._build_filter_tuples(system_filter_exceptions)

            for i, item in enumerate(filter_tuples):
                ws2.cell(row=3 + i, column=1, value=f"规则{i+1}: {self._format_filter_item(item)}")

            offset = len(filter_tuples)
            for j, item in enumerate(exception_tuples):
                ws2.cell(
                    row=3 + offset + j,
                    column=1,
                    value=f"例外{j+1}: {self._format_filter_item(item)}"
                )

            # 应用筛选条件（主筛选AND + 例外保留OR）
            df_filtered = CompareEngine.apply_filters(
                df_filtered,
                filter_tuples,
                exception_tuples
            )
            
            line_count = len(filter_tuples) + len(exception_tuples) + 1
            ws2.cell(row=2 + line_count, column=1, value=f"筛选后剩余 {len(df_filtered)} 行")
            start_row = 4 + line_count
        else:
            ws2.cell(row=3, column=1, value="（无筛选规则）")
            start_row = 6
        
        for c_idx, col in enumerate(df_filtered.columns, 1):
            cell = ws2.cell(row=start_row, column=c_idx, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")
            cell.border = thin_border
        for r_idx, row in enumerate(df_filtered.itertuples(index=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=safe_value(value))
                cell.border = thin_border
        
        # === Sheet3: 透视计算结果 ===
        ws3 = wb.create_sheet("3-透视计算结果")
        pivot_config = config.get("pivot_column", {})
        pivot_col = pivot_config.get("system") if isinstance(pivot_config, dict) else pivot_config
        pivot_values = config.get("pivot_values", [])
        if not isinstance(pivot_values, (list, tuple)):
            pivot_values = []
        
        ws3.cell(row=1, column=1, value="【系统表透视配置】").font = Font(bold=True, size=12, color="2E7D32")
        
        if pivot_col:
            ws3.cell(row=2, column=1, value=f"透视列: {pivot_col}")
            ws3.cell(row=3, column=1, value=f"透视值: {', '.join(map(str, pivot_values)) if pivot_values else '(全部)'}")
            
            # 执行透视计算
            key_mappings = config.get("key_mappings", [])
            system_key_cols = [
                (k.get("system") or k.get("system_col"))
                for k in key_mappings
                if isinstance(k, dict) and (k.get("system") or k.get("system_col"))
            ]
            value_mapping = config.get("value_mapping", {})
            system_val_col = value_mapping.get("system", "") or value_mapping.get("system_col", "")
            
            if system_key_cols and system_val_col:
                system_with_key = CompareEngine.make_key(df_filtered, system_key_cols)
                
                try:
                    # 使用 aggregate_data 执行透视
                    pivot_df, detected_pivot_values = CompareEngine.aggregate_data(
                        system_with_key, "__KEY__", [system_val_col], pivot_col
                    )
                    
                    start_row = 5
                    for c_idx, col in enumerate(pivot_df.columns, 1):
                        cell = ws3.cell(row=start_row, column=c_idx, value=col)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        if col == "__KEY__":
                            cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid")
                        elif col == "系统总计":
                            cell.fill = PatternFill(start_color="FFF3E0", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
                    
                    for r_idx, row in enumerate(pivot_df.itertuples(index=False), start_row + 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                            cell.border = thin_border
                            
                    ws3.cell(row=start_row + len(pivot_df) + 2, column=1, 
                             value=f"透视后共 {len(pivot_df)} 行").font = Font(italic=True)
                except Exception as e:
                    ws3.cell(row=5, column=1, value=f"透视计算出错: {str(e)}")
            else:
                ws3.cell(row=5, column=1, value="（请先配置主键和数值列）")
        else:
            ws3.cell(row=2, column=1, value="（未配置系统表透视列）")
            
            # 如果没有透视，显示按主键汇总的结果
            key_mappings = config.get("key_mappings", [])
            system_key_cols = [
                (k.get("system") or k.get("system_col"))
                for k in key_mappings
                if isinstance(k, dict) and (k.get("system") or k.get("system_col"))
            ]
            value_mapping = config.get("value_mapping", {})
            system_val_col = value_mapping.get("system", "") or value_mapping.get("system_col", "")
            
            if system_key_cols and system_val_col:
                system_with_key = CompareEngine.make_key(df_filtered, system_key_cols)
                
                # 按主键汇总
                agg_df = system_with_key.groupby("__KEY__", as_index=False)[system_val_col].sum()
                agg_df.rename(columns={system_val_col: "系统总计"}, inplace=True)
                
                ws3.cell(row=3, column=1, value="（按主键汇总）")
                start_row = 5
                for c_idx, col in enumerate(agg_df.columns, 1):
                    cell = ws3.cell(row=start_row, column=c_idx, value=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
                    cell.border = thin_border
                
                for r_idx, row in enumerate(agg_df.itertuples(index=False), start_row + 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                        cell.border = thin_border
        
        # 保存文件
        from PyQt6.QtWidgets import QFileDialog
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"系统表预处理预览_{timestamp}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出系统表预处理预览", default_name, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb.save(file_path)
                show_info(self, "导出成功", f"系统表预处理预览已保存:\n{file_path}")
                os.startfile(file_path)  # 自动打开
            except Exception as e:
                show_warning(self, "导出失败", f"保存文件时出错: {str(e)}")

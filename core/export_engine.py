"""
导出引擎 - 生成带颜色的 Excel 文件
"""
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from config import EXCEL_COLORS, COMPARE_STATUS


class ExportEngine:
    """Excel 导出引擎"""

    @staticmethod
    def create_fill(color_key: str) -> PatternFill:
        """创建填充样式"""
        color = EXCEL_COLORS.get(color_key, "FFFFFFFF")
        return PatternFill(fill_type="solid", start_color=color, end_color=color)

    @staticmethod
    def export_results(
        out_path: str,
        result_df: pd.DataFrame,
        pivot_values: List[str],
        config_info: Dict[str, Any]
    ):
        """
        导出比对结果到 Excel
        
        Args:
            out_path: 输出文件路径
            result_df: 比对结果 DataFrame
            pivot_values: 透视值列表
            config_info: 配置信息字典
        """
        wb = Workbook()
        
        # --- Sheet 1: 完整结果 ---
        ws_all = wb.active
        ws_all.title = "📋 完整结果"
        
        # 准备导出列
        export_cols = ExportEngine._get_export_columns(result_df, pivot_values)
        available_cols = [c for c in export_cols if c in result_df.columns]
        export_df = result_df[available_cols].copy() if available_cols else result_df.copy()

        # 排序，确保完整结果表稳定可核对
        if "__KEY__" in export_df.columns:
            export_df = export_df.sort_values(by="__KEY__", kind="stable").reset_index(drop=True)

        # 统一缺失值处理：状态和主键保留文本，其余列空值按0填充便于核对
        for col in export_df.columns:
            if col in ["__KEY__", "比对状态"]:
                export_df[col] = export_df[col].fillna("")
                continue
            if pd.api.types.is_numeric_dtype(export_df[col]):
                export_df[col] = export_df[col].fillna(0)
        
        # 写入数据
        ExportEngine._write_dataframe(ws_all, export_df)
        
        # 应用颜色
        ExportEngine._apply_colors(ws_all, export_df)
        
        # 自动列宽
        ExportEngine._auto_width(ws_all)
        
        # --- Sheet 2: 仅差异 ---
        diff_df = export_df[export_df["比对状态"] != COMPARE_STATUS["match"]].copy()
        if not diff_df.empty:
            ws_diff = wb.create_sheet(title="📌 差异数据")
            ExportEngine._write_dataframe(ws_diff, diff_df)
            ExportEngine._apply_colors(ws_diff, diff_df)
            ExportEngine._auto_width(ws_diff)
        
        # --- Sheet 3: 说明 ---
        ws_meta = wb.create_sheet(title="ℹ️ 说明")
        ExportEngine._write_metadata(ws_meta, result_df, diff_df, config_info, pivot_values)
        ExportEngine._auto_width(ws_meta)
        
        # 保存
        wb.save(out_path)

    @staticmethod
    def _get_export_columns(df: pd.DataFrame, pivot_values: List[str]) -> List[str]:
        """
        获取要导出的列（严格顺序，排除中间列）
        
        列顺序规范（v1.4.1）：
        1. 主键 (__KEY__)
        2. 透视列（按字母排序）
        3. 系统总计
        4. 手工数量
        5. 差值
        6. 比对状态
        
        注意：此顺序需与 qt_result_preview.py 中的 _get_export_columns 保持一致
        """
        cols = []
        
        # 1. 主键
        if "__KEY__" in df.columns:
            cols.append("__KEY__")
        
        # 2. 透视列（按排序）
        for pv in sorted(pivot_values):
            if pv in df.columns:
                cols.append(pv)
        
        # 3. 系统总计
        if "系统总计" in df.columns:
            cols.append("系统总计")
        
        # 4. 手工数量（移到系统总计之后）
        if "手工数量" in df.columns:
            cols.append("手工数量")
        
        # 5. 差值
        if "差值" in df.columns:
            cols.append("差值")
        
        # 6. 比对状态
        if "比对状态" in df.columns:
            cols.append("比对状态")
        
        # 排除带后缀的中间列
        exclude_suffixes = ('_manual', '_system', '_x', '_y', '_left', '_right')
        final_cols = [c for c in cols if not any(c.endswith(suffix) for suffix in exclude_suffixes)]
        
        return final_cols if final_cols else list(df.columns)

    @staticmethod
    def _write_dataframe(ws, df: pd.DataFrame):
        """写入 DataFrame 到工作表"""
        # 写入表头
        header_fill = ExportEngine.create_fill("header")
        header_font = Font(bold=True)
        
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 处理 NaN
                if pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = value

    @staticmethod
    def _apply_colors(ws, df: pd.DataFrame):
        """应用行颜色"""
        if "比对状态" not in df.columns:
            return
        
        status_idx = list(df.columns).index("比对状态")
        diff_idx = list(df.columns).index("差值") if "差值" in df.columns else None
        
        for row_idx in range(2, len(df) + 2):
            status = ws.cell(row=row_idx, column=status_idx + 1).value
            diff_val = ws.cell(row=row_idx, column=diff_idx + 1).value if diff_idx is not None else 0
            
            # 确定颜色
            fill = None
            if status == COMPARE_STATUS["match"]:
                fill = ExportEngine.create_fill("match")
            elif status == COMPARE_STATUS["diff"]:
                try:
                    diff_num = float(diff_val) if diff_val else 0
                    fill = ExportEngine.create_fill("diff_pos" if diff_num > 0 else "diff_neg")
                except (ValueError, TypeError):
                    fill = ExportEngine.create_fill("diff_pos")
            elif status in [COMPARE_STATUS["system_only"], COMPARE_STATUS["manual_only"]]:
                fill = ExportEngine.create_fill("missing")
            
            # 应用颜色到整行
            if fill:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    @staticmethod
    def _auto_width(ws):
        """自动调整列宽"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    # 中文字符宽度调整
                    chinese_count = sum(1 for c in str(cell.value or '') if '\u4e00' <= c <= '\u9fff')
                    cell_len += chinese_count * 0.5
                    max_len = max(max_len, cell_len)
                except:
                    pass
            
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    @staticmethod
    def _write_metadata(
        ws,
        result_df: pd.DataFrame,
        diff_df: pd.DataFrame,
        config_info: Dict[str, Any],
        pivot_values: List[str]
    ):
        """写入元数据说明"""
        # 处理透视列配置（可能是字典或字符串）
        pivot_col = config_info.get("pivot_column", "")
        if isinstance(pivot_col, dict):
            pivot_col = pivot_col.get("system", "未使用")
        pivot_col = pivot_col if pivot_col else "未使用"
        
        # 处理主键字段（可能是列表）
        key_columns = config_info.get("key_columns", "")
        if isinstance(key_columns, list):
            key_columns = ", ".join(str(k) for k in key_columns)
        elif isinstance(key_columns, dict):
            key_columns = str(key_columns)
        
        # 处理数值字段
        value_columns = config_info.get("value_columns", "")
        if isinstance(value_columns, dict):
            value_columns = f"手工: {value_columns.get('manual', '')}, 系统: {value_columns.get('system', '')}"
        
        data = [
            ["📊 对账结果导出"],
            [],
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [],
            ["【统计结果】", ""],
            ["总记录数", len(result_df)],
            ["✓ 完全匹配", len(result_df[result_df["比对状态"] == COMPARE_STATUS["match"]]) if "比对状态" in result_df.columns else 0],
            ["↕ 数量差异", len(result_df[result_df["比对状态"] == COMPARE_STATUS["diff"]]) if "比对状态" in result_df.columns else 0],
            ["✗ 系统缺失", len(result_df[result_df["比对状态"] == COMPARE_STATUS["manual_only"]]) if "比对状态" in result_df.columns else 0],
            ["✗ 手工缺失", len(result_df[result_df["比对状态"] == COMPARE_STATUS["system_only"]]) if "比对状态" in result_df.columns else 0],
            [],
            ["【配置信息】", ""],
            ["主键字段", key_columns],
            ["数值字段", value_columns],
            ["透视列", pivot_col],
            ["差值公式", config_info.get("difference_formula", config_info.get("diff_formula", "手工数量 - 系统总计"))],
            [],
            ["【颜色说明】", ""],
            ["绿色", "✓ 一致"],
            ["浅黄绿", "↑ 手工多（差值>0）"],
            ["浅蓝", "↓ 手工少（差值<0）"],
            ["浅红", "✗ 缺失（单边存在）"],
        ]
        
        if pivot_values:
            data.append([])
            data.append(["【透视值】", ", ".join(pivot_values)])
        
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# 将项目根目录加入路径，便于直接导入 core
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.compare_engine import CompareEngine

# 输入文件
MANUAL_FILE = Path("222.xlsx")
SYSTEM_FILE = Path("送货单执行报表 (73).xlsx")

# 输出目录（用户可指定，默认当前目录）
OUT_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 输出文件
MANUAL_OUT = OUT_DIR / "手工表预处理预览_ASN_test.xlsx"
SYSTEM_OUT = OUT_DIR / "系统表预处理预览_ASN_test.xlsx"

# 模板文件（APPDATA 路径）
TEMPLATE_PATH = Path.home() / "AppData" / "Roaming" / "SupplyChain-Reconciler" / "templates.json"


def load_asn_template():
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_PATH}")
    data = json.loads(TEMPLATE_PATH.read_text(encoding="utf-8"))
    for item in data:
        if item.get("name") == "ASN":
            return item.get("config", {})
    raise ValueError("未找到名为 ASN 的模板")


def apply_filters(df: pd.DataFrame, filters):
    """按 aggregate_data 的规则应用筛选"""
    if not filters:
        return df
    df = df.copy()
    for col, op, val in filters:
        if col not in df.columns:
            continue
        col_data = df[col].astype(str)
        if op == "EQUALS":
            df = df[col_data == val]
        elif op == "NOT_EQUALS":
            df = df[col_data != val]
        elif op == "CONTAINS":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            if values:
                mask = col_data.str.contains(values[0], na=False, regex=False)
                for v in values[1:]:
                    mask |= col_data.str.contains(v, na=False, regex=False)
                df = df[mask]
        elif op == "NOT_CONTAINS":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            if values:
                mask = ~col_data.str.contains(values[0], na=False, regex=False)
                for v in values[1:]:
                    mask &= ~col_data.str.contains(v, na=False, regex=False)
                df = df[mask]
        elif op == "IN_LIST":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            df = df[col_data.isin(values)]
        elif op == "NOT_IN_LIST":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            df = df[~col_data.isin(values)]
    return df


def export_manual(template):
    key_mappings = template.get("key_mappings", [])
    value_mapping = template.get("value_mapping", {})
    manual_key_cols = [m["manual"] for m in key_mappings if m.get("manual")]
    manual_val_col = value_mapping.get("manual", "")
    manual_filters = [(f["column"], f["operator"], f["value"]) for f in template.get("manual_filters", [])]
    manual_pivot = template.get("manual_pivot", {})
    clean_rules = template.get("clean_rules", [])

    df_original = pd.read_excel(MANUAL_FILE)

    # 清洗
    df_cleaned = CompareEngine.clean_column(df_original.copy(), clean_rules) if clean_rules else df_original.copy()

    # 透视计算
    pivot_df = None
    out_cols = []
    in_cols = []
    if manual_key_cols and manual_val_col:
        manual_with_key = CompareEngine.make_key(df_cleaned, manual_key_cols)
        pivot_df, out_cols, in_cols = CompareEngine.aggregate_manual_with_pivot(
            manual_with_key,
            "__KEY__",
            manual_val_col,
            manual_pivot,
            manual_filters
        )

    # 使用 openpyxl 创建带格式的 Excel
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def safe_value(val):
        if pd.isna(val) or val is pd.NA:
            return ""
        if isinstance(val, (np.integer, np.floating)):
            return float(val) if np.isfinite(val) else ""
        return val
    
    # === Sheet1: 原始数据 ===
    ws1 = wb.active
    ws1.title = "1-原始数据"
    ws1.cell(row=1, column=1, value="【手工表原始数据】").font = Font(bold=True, size=12, color="0000FF")
    ws1.cell(row=2, column=1, value=f"共 {len(df_original)} 行数据")
    
    start_row = 4
    for c_idx, col in enumerate(df_original.columns, 1):
        cell = ws1.cell(row=start_row, column=c_idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")
        cell.border = thin_border
    for r_idx, row in enumerate(df_original.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=safe_value(value))
            cell.border = thin_border
    
    # === Sheet2: 清洗后数据 ===
    ws2 = wb.create_sheet("2-清洗后数据")
    ws2.cell(row=1, column=1, value="【清洗规则】").font = Font(bold=True, size=12, color="FF0000")
    
    if clean_rules:
        for i, rule in enumerate(clean_rules):
            mode_text = f"{rule['column']}: {rule['mode']} 正则'{rule.get('regexes', [])}'"
            if rule.get('replace'):
                mode_text += f" => '{rule['replace']}'"
            ws2.cell(row=2+i, column=1, value=f"规则{i+1}: {mode_text}")
        start_row = 4 + len(clean_rules)
    else:
        ws2.cell(row=2, column=1, value="（无清洗规则）")
        start_row = 4
    
    for c_idx, col in enumerate(df_cleaned.columns, 1):
        cell = ws2.cell(row=start_row, column=c_idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
        cell.border = thin_border
    for r_idx, row in enumerate(df_cleaned.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=safe_value(value))
            cell.border = thin_border
    
    # === Sheet3: 透视计算结果 ===
    ws3 = wb.create_sheet("3-透视计算结果")
    ws3.cell(row=1, column=1, value="【手工表透视配置】").font = Font(bold=True, size=12, color="0000FF")
    
    if manual_pivot and manual_pivot.get("pivot_column"):
        pivot_col = manual_pivot.get("pivot_column", "")
        out_values = manual_pivot.get("out_values", [])
        in_values = manual_pivot.get("in_values", [])
        
        ws3.cell(row=2, column=1, value=f"透视列: {pivot_col}")
        ws3.cell(row=3, column=1, value=f"📤 出库值: {', '.join(out_values) if out_values else '(无)'}")
        ws3.cell(row=4, column=1, value=f"📥 入库值: {', '.join(in_values) if in_values else '(无)'}")
        ws3.cell(row=5, column=1, value="计算公式: 手工数量 = Σ出库 - Σ入库").font = Font(bold=True, color="FF6600")
        
        if pivot_df is not None:
            start_row = 7
            for c_idx, col in enumerate(pivot_df.columns, 1):
                cell = ws3.cell(row=start_row, column=c_idx, value=col)
                cell.font = Font(bold=True)
                cell.border = thin_border
                if col in out_values:
                    cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")  # 蓝色-出库
                elif col in in_values:
                    cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")  # 绿色-入库
                elif col == "手工数量":
                    cell.fill = PatternFill(start_color="FFF3E0", fill_type="solid")  # 橙色-结果
            
            for r_idx, row in enumerate(pivot_df.itertuples(index=False), start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                    cell.border = thin_border
        else:
            ws3.cell(row=7, column=1, value="（请先配置主键和数值列）")
    else:
        ws3.cell(row=2, column=1, value="（未配置手工表透视）")
    
    wb.save(MANUAL_OUT)
    print(f"手工表预处理导出完成: {MANUAL_OUT}")


def export_system(template):
    key_mappings = template.get("key_mappings", [])
    value_mapping = template.get("value_mapping", {})
    system_key_cols = [m["system"] for m in key_mappings if m.get("system")]
    system_val_col = value_mapping.get("system", "")
    system_filters = [(f["column"], f["operator"], f["value"]) for f in template.get("system_filters", [])]
    pivot_col_cfg = template.get("pivot_column", {})
    pivot_col = pivot_col_cfg.get("system") if isinstance(pivot_col_cfg, dict) else pivot_col_cfg
    pivot_values = template.get("pivot_values", [])

    df_original = pd.read_excel(SYSTEM_FILE)

    # 筛选
    df_filtered = apply_filters(df_original, system_filters)

    pivot_df = None
    if system_key_cols and system_val_col:
        system_with_key = CompareEngine.make_key(df_filtered, system_key_cols)
        pivot_df, _ = CompareEngine.aggregate_data(
            system_with_key,
            "__KEY__",
            [system_val_col],
            pivot_col,
            system_filters
        )

    # 使用 openpyxl 创建带格式的 Excel
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def safe_value(val):
        if pd.isna(val) or val is pd.NA:
            return ""
        if isinstance(val, (np.integer, np.floating)):
            return float(val) if np.isfinite(val) else ""
        return val
    
    # === Sheet1: 原始数据 ===
    ws1 = wb.active
    ws1.title = "1-原始数据"
    ws1.cell(row=1, column=1, value="【系统表原始数据】").font = Font(bold=True, size=12, color="0000FF")
    ws1.cell(row=2, column=1, value=f"共 {len(df_original)} 行数据")
    
    start_row = 4
    for c_idx, col in enumerate(df_original.columns, 1):
        cell = ws1.cell(row=start_row, column=c_idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")
        cell.border = thin_border
    for r_idx, row in enumerate(df_original.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=safe_value(value))
            cell.border = thin_border
    
    # === Sheet2: 筛选后数据 ===
    ws2 = wb.create_sheet("2-筛选后数据")
    ws2.cell(row=1, column=1, value="【筛选条件】").font = Font(bold=True, size=12, color="FF0000")
    
    if system_filters:
        for i, (col, op, val) in enumerate(system_filters):
            ws2.cell(row=2+i, column=1, value=f"条件{i+1}: {col} {op} '{val}'")
        start_row = 4 + len(system_filters)
    else:
        ws2.cell(row=2, column=1, value="（无筛选条件）")
        start_row = 4
    
    ws2.cell(row=start_row-1, column=1, value=f"共 {len(df_filtered)} 行数据（筛选后）")
    
    for c_idx, col in enumerate(df_filtered.columns, 1):
        cell = ws2.cell(row=start_row, column=c_idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
        cell.border = thin_border
    for r_idx, row in enumerate(df_filtered.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=safe_value(value))
            cell.border = thin_border
    
    # === Sheet3: 透视计算结果 ===
    ws3 = wb.create_sheet("3-透视计算结果")
    ws3.cell(row=1, column=1, value="【系统表透视配置】").font = Font(bold=True, size=12, color="0000FF")
    
    if pivot_col:
        ws3.cell(row=2, column=1, value=f"透视列: {pivot_col}")
        ws3.cell(row=3, column=1, value=f"透视值: {', '.join(pivot_values) if pivot_values else '(全部)'}")
        ws3.cell(row=4, column=1, value=f"数值列: {system_val_col}")
        ws3.cell(row=5, column=1, value="计算公式: 系统总计 = Σ各透视列").font = Font(bold=True, color="FF6600")
        
        if pivot_df is not None:
            start_row = 7
            for c_idx, col in enumerate(pivot_df.columns, 1):
                cell = ws3.cell(row=start_row, column=c_idx, value=col)
                cell.font = Font(bold=True)
                cell.border = thin_border
                if col in pivot_values:
                    cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")  # 蓝色-透视列
                elif col == "系统总计":
                    cell.fill = PatternFill(start_color="FFF3E0", fill_type="solid")  # 橙色-结果
            
            for r_idx, row in enumerate(pivot_df.itertuples(index=False), start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                    cell.border = thin_border
        else:
            ws3.cell(row=7, column=1, value="（请先配置主键和数值列）")
    else:
        ws3.cell(row=2, column=1, value="（未配置系统表透视，直接汇总）")
        ws3.cell(row=3, column=1, value=f"数值列: {system_val_col}")
        
        if pivot_df is not None:
            start_row = 5
            for c_idx, col in enumerate(pivot_df.columns, 1):
                cell = ws3.cell(row=start_row, column=c_idx, value=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E3F2FD", fill_type="solid")
                cell.border = thin_border
            
            for r_idx, row in enumerate(pivot_df.itertuples(index=False), start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=safe_value(value))
                    cell.border = thin_border
    
    wb.save(SYSTEM_OUT)
    print(f"系统表预处理导出完成: {SYSTEM_OUT}")


def main():
    template = load_asn_template()
    export_manual(template)
    export_system(template)


if __name__ == "__main__":
    main()

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 将项目根目录加入路径，便于直接导入 core
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.compare_engine import CompareEngine

# 输入文件
MANUAL_FILE = Path("222.xlsx")
SYSTEM_FILE = Path("送货单执行报表 (73).xlsx")

# 输出目录（用户可指定，默认当前目录）
OUT_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 输出文件
MANUAL_OUT = OUT_DIR / "手工表预处理预览_ASN_test.xlsx"
SYSTEM_OUT = OUT_DIR / "系统表预处理预览_ASN_test.xlsx"

# 模板文件（APPDATA 路径）
TEMPLATE_PATH = Path.home() / "AppData" / "Roaming" / "SupplyChain-Reconciler" / "templates.json"


def load_asn_template():
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_PATH}")
    data = json.loads(TEMPLATE_PATH.read_text(encoding="utf-8"))
    for item in data:
        if item.get("name") == "ASN":
            return item.get("config", {})
    raise ValueError("未找到名为 ASN 的模板")


def apply_filters(df: pd.DataFrame, filters):
    """按 aggregate_data 的规则应用筛选"""
    if not filters:
        return df
    df = df.copy()
    for col, op, val in filters:
        if col not in df.columns:
            continue
        col_data = df[col].astype(str)
        if op == "EQUALS":
            df = df[col_data == val]
        elif op == "NOT_EQUALS":
            df = df[col_data != val]
        elif op == "CONTAINS":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            if values:
                mask = col_data.str.contains(values[0], na=False, regex=False)
                for v in values[1:]:
                    mask |= col_data.str.contains(v, na=False, regex=False)
                df = df[mask]
        elif op == "NOT_CONTAINS":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            if values:
                mask = ~col_data.str.contains(values[0], na=False, regex=False)
                for v in values[1:]:
                    mask &= ~col_data.str.contains(v, na=False, regex=False)
                df = df[mask]
        elif op == "IN_LIST":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            df = df[col_data.isin(values)]
        elif op == "NOT_IN_LIST":
            values = [v.strip() for v in str(val).replace('；', ';').replace('，', ',').replace(';', ',').split(',') if v.strip()]
            df = df[~col_data.isin(values)]
    return df


def export_manual(template):
    key_mappings = template.get("key_mappings", [])
    value_mapping = template.get("value_mapping", {})
    manual_key_cols = [m["manual"] for m in key_mappings if m.get("manual")]
    manual_val_col = value_mapping.get("manual", "")
    manual_filters = [(f["column"], f["operator"], f["value"]) for f in template.get("manual_filters", [])]
    manual_pivot = template.get("manual_pivot", {})

    df_original = pd.read_excel(MANUAL_FILE)

    # 清洗
    clean_rules = template.get("clean_rules", [])
    df_cleaned = CompareEngine.clean_column(df_original.copy(), clean_rules) if clean_rules else df_original.copy()

    # 透视计算
    pivot_df = None
    if manual_key_cols and manual_val_col:
        manual_with_key = CompareEngine.make_key(df_cleaned, manual_key_cols)
        pivot_df, _, _ = CompareEngine.aggregate_manual_with_pivot(
            manual_with_key,
            "__KEY__",
            manual_val_col,
            manual_pivot,
            manual_filters
        )

    with pd.ExcelWriter(MANUAL_OUT, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name="1-原始数据", index=False)
        df_cleaned.to_excel(writer, sheet_name="2-清洗后数据", index=False)
        if pivot_df is not None:
            pivot_df.to_excel(writer, sheet_name="3-透视计算结果", index=False)

    # 追加透视配置说明 Sheet，方便核对
    try:
        wb = load_workbook(MANUAL_OUT)
        ws = wb.create_sheet("透视配置")
        ws.cell(1, 1, "【手工表透视配置】").font = Font(bold=True, color="1565C0", size=12)
        if manual_pivot:
            ws.cell(2, 1, f"透视列: {manual_pivot.get('pivot_column', '')}")
            ws.cell(3, 1, f"出库值: {', '.join(manual_pivot.get('out_values', [])) or '(无)'}")
            ws.cell(4, 1, f"入库值: {', '.join(manual_pivot.get('in_values', [])) or '(无)'}")
            ws.cell(5, 1, "计算公式: 手工数量 = Σ出库 - Σ入库").font = Font(color="D35400")
        wb.save(MANUAL_OUT)
    except Exception as e:
        print(f"附加手工透视配置说明失败: {e}")

    print(f"手工表预处理导出完成: {MANUAL_OUT}")


def export_system(template):
    key_mappings = template.get("key_mappings", [])
    value_mapping = template.get("value_mapping", {})
    system_key_cols = [m["system"] for m in key_mappings if m.get("system")]
    system_val_col = value_mapping.get("system", "")
    system_filters = [(f["column"], f["operator"], f["value"]) for f in template.get("system_filters", [])]
    pivot_col_cfg = template.get("pivot_column", {})
    pivot_col = pivot_col_cfg.get("system") if isinstance(pivot_col_cfg, dict) else pivot_col_cfg

    df_original = pd.read_excel(SYSTEM_FILE)

    # 筛选
    df_filtered = apply_filters(df_original, system_filters)

    pivot_df = None
    if system_key_cols and system_val_col:
        system_with_key = CompareEngine.make_key(df_filtered, system_key_cols)
        pivot_df, _ = CompareEngine.aggregate_data(
            system_with_key,
            "__KEY__",
            [system_val_col],
            pivot_col,
            system_filters
        )

    with pd.ExcelWriter(SYSTEM_OUT, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name="1-原始数据", index=False)
        df_filtered.to_excel(writer, sheet_name="2-筛选后数据", index=False)
        if pivot_df is not None:
            pivot_df.to_excel(writer, sheet_name="3-透视计算结果", index=False)

    print(f"系统表预处理导出完成: {SYSTEM_OUT}")


def main():
    template = load_asn_template()
    export_manual(template)
    export_system(template)


if __name__ == "__main__":
    main()

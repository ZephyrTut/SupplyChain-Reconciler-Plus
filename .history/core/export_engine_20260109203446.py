"""
å¯¼å‡ºå¼•æ“ - ç”Ÿæˆå¸¦é¢œè‰²çš„ Excel æ–‡ä»¶
"""
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from config import EXCEL_COLORS, COMPARE_STATUS


class ExportEngine:
    """Excel å¯¼å‡ºå¼•æ“"""

    @staticmethod
    def create_fill(color_key: str) -> PatternFill:
        """åˆ›å»ºå¡«å……æ ·å¼"""
        color = EXCEL_COLORS.get(color_key, "FFFFFFFF")
        return PatternFill(fill_type="solid", start_color=color, end_color=color)

    @staticmethod
    def export_results(
        out_path: str,
        result_df: pd.DataFrame,
        pivot_values: List[str],
        config_info: Dict[str, Any]
    ):
        """
        å¯¼å‡ºæ¯”å¯¹ç»“æœåˆ° Excel
        
        Args:
            out_path: è¾“å‡ºæ–‡ä»¶è·¯å¾„
            result_df: æ¯”å¯¹ç»“æœ DataFrame
            pivot_values: é€è§†å€¼åˆ—è¡¨
            config_info: é…ç½®ä¿¡æ¯å­—å…¸
        """
        wb = Workbook()
        
        # --- Sheet 1: å®Œæ•´ç»“æœ ---
        ws_all = wb.active
        ws_all.title = "ğŸ“‹ å®Œæ•´ç»“æœ"
        
        # å‡†å¤‡å¯¼å‡ºåˆ—
        export_cols = ExportEngine._get_export_columns(result_df, pivot_values)
        export_df = result_df[export_cols].copy() if all(c in result_df.columns for c in export_cols) else result_df.copy()
        
        # å†™å…¥æ•°æ®
        ExportEngine._write_dataframe(ws_all, export_df)
        
        # åº”ç”¨é¢œè‰²
        ExportEngine._apply_colors(ws_all, export_df)
        
        # è‡ªåŠ¨åˆ—å®½
        ExportEngine._auto_width(ws_all)
        
        # --- Sheet 2: ä»…å·®å¼‚ ---
        diff_df = export_df[export_df["æ¯”å¯¹çŠ¶æ€"] != COMPARE_STATUS["match"]].copy()
        if not diff_df.empty:
            ws_diff = wb.create_sheet(title="ğŸ“Œ å·®å¼‚æ•°æ®")
            ExportEngine._write_dataframe(ws_diff, diff_df)
            ExportEngine._apply_colors(ws_diff, diff_df)
            ExportEngine._auto_width(ws_diff)
        
        # --- Sheet 3: è¯´æ˜ ---
        ws_meta = wb.create_sheet(title="â„¹ï¸ è¯´æ˜")
        ExportEngine._write_metadata(ws_meta, result_df, diff_df, config_info, pivot_values)
        ExportEngine._auto_width(ws_meta)
        
        # ä¿å­˜
        wb.save(out_path)

    @staticmethod
    def _get_export_columns(df: pd.DataFrame, pivot_values: List[str]) -> List[str]:
        """
        è·å–è¦å¯¼å‡ºçš„åˆ—ï¼ˆä¸¥æ ¼é¡ºåºï¼Œæ’é™¤ä¸­é—´åˆ—ï¼‰
        
        åˆ—é¡ºåºè§„èŒƒï¼ˆv1.4.1ï¼‰ï¼š
        1. ä¸»é”® (__KEY__)
        2. é€è§†åˆ—ï¼ˆæŒ‰å­—æ¯æ’åºï¼‰
        3. ç³»ç»Ÿæ€»è®¡
        4. æ‰‹å·¥æ•°é‡
        5. å·®å€¼
        6. æ¯”å¯¹çŠ¶æ€
        
        æ³¨æ„ï¼šæ­¤é¡ºåºéœ€ä¸ qt_result_preview.py ä¸­çš„ _get_export_columns ä¿æŒä¸€è‡´
        """
        cols = []
        
        # 1. ä¸»é”®
        if "__KEY__" in df.columns:
            cols.append("__KEY__")
        
        # 2. é€è§†åˆ—ï¼ˆæŒ‰æ’åºï¼‰
        for pv in sorted(pivot_values):
            if pv in df.columns:
                cols.append(pv)
        
        # 3. ç³»ç»Ÿæ€»è®¡
        if "ç³»ç»Ÿæ€»è®¡" in df.columns:
            cols.append("ç³»ç»Ÿæ€»è®¡")
        
        # 4. æ‰‹å·¥æ•°é‡ï¼ˆç§»åˆ°ç³»ç»Ÿæ€»è®¡ä¹‹åï¼‰
        if "æ‰‹å·¥æ•°é‡" in df.columns:
            cols.append("æ‰‹å·¥æ•°é‡")
        
        # 5. å·®å€¼
        if "å·®å€¼" in df.columns:
            cols.append("å·®å€¼")
        
        # 6. æ¯”å¯¹çŠ¶æ€
        if "æ¯”å¯¹çŠ¶æ€" in df.columns:
            cols.append("æ¯”å¯¹çŠ¶æ€")
        
        # æ’é™¤å¸¦åç¼€çš„ä¸­é—´åˆ—
        exclude_suffixes = ('_manual', '_system', '_x', '_y', '_left', '_right')
        final_cols = [c for c in cols if not any(c.endswith(suffix) for suffix in exclude_suffixes)]
        
        return final_cols if final_cols else list(df.columns)

    @staticmethod
    def _write_dataframe(ws, df: pd.DataFrame):
        """å†™å…¥ DataFrame åˆ°å·¥ä½œè¡¨"""
        # å†™å…¥è¡¨å¤´
        header_fill = ExportEngine.create_fill("header")
        header_font = Font(bold=True)
        
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # å†™å…¥æ•°æ®
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # å¤„ç† NaN
                if pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = value

    @staticmethod
    def _apply_colors(ws, df: pd.DataFrame):
        """åº”ç”¨è¡Œé¢œè‰²"""
        if "æ¯”å¯¹çŠ¶æ€" not in df.columns:
            return
        
        status_idx = list(df.columns).index("æ¯”å¯¹çŠ¶æ€")
        diff_idx = list(df.columns).index("å·®å€¼") if "å·®å€¼" in df.columns else None
        
        for row_idx in range(2, len(df) + 2):
            status = ws.cell(row=row_idx, column=status_idx + 1).value
            diff_val = ws.cell(row=row_idx, column=diff_idx + 1).value if diff_idx is not None else 0
            
            # ç¡®å®šé¢œè‰²
            fill = None
            if status == COMPARE_STATUS["match"]:
                fill = ExportEngine.create_fill("match")
            elif status == COMPARE_STATUS["diff"]:
                try:
                    diff_num = float(diff_val) if diff_val else 0
                    fill = ExportEngine.create_fill("diff_pos" if diff_num > 0 else "diff_neg")
                except (ValueError, TypeError):
                    fill = ExportEngine.create_fill("diff_pos")
            elif status in [COMPARE_STATUS["system_only"], COMPARE_STATUS["manual_only"]]:
                fill = ExportEngine.create_fill("missing")
            
            # åº”ç”¨é¢œè‰²åˆ°æ•´è¡Œ
            if fill:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    @staticmethod
    def _auto_width(ws):
        """è‡ªåŠ¨è°ƒæ•´åˆ—å®½"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    # ä¸­æ–‡å­—ç¬¦å®½åº¦è°ƒæ•´
                    chinese_count = sum(1 for c in str(cell.value or '') if '\u4e00' <= c <= '\u9fff')
                    cell_len += chinese_count * 0.5
                    max_len = max(max_len, cell_len)
                except:
                    pass
            
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    @staticmethod
    def _write_metadata(
        ws,
        result_df: pd.DataFrame,
        diff_df: pd.DataFrame,
        config_info: Dict[str, Any],
        pivot_values: List[str]
    ):
        """å†™å…¥å…ƒæ•°æ®è¯´æ˜"""
        data = [
            ["ğŸ“Š å¯¹è´¦ç»“æœå¯¼å‡º"],
            [],
            ["å¯¼å‡ºæ—¶é—´", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [],
            ["ã€ç»Ÿè®¡ç»“æœã€‘", ""],
            ["æ€»è®°å½•æ•°", len(result_df)],
            ["âœ“ å®Œå…¨åŒ¹é…", len(result_df[result_df["æ¯”å¯¹çŠ¶æ€"] == COMPARE_STATUS["match"]]) if "æ¯”å¯¹çŠ¶æ€" in result_df.columns else 0],
            ["â†• æ•°é‡å·®å¼‚", len(result_df[result_df["æ¯”å¯¹çŠ¶æ€"] == COMPARE_STATUS["diff"]]) if "æ¯”å¯¹çŠ¶æ€" in result_df.columns else 0],
            ["âœ— ç³»ç»Ÿç¼ºå¤±", len(result_df[result_df["æ¯”å¯¹çŠ¶æ€"] == COMPARE_STATUS["manual_only"]]) if "æ¯”å¯¹çŠ¶æ€" in result_df.columns else 0],
            ["âœ— æ‰‹å·¥ç¼ºå¤±", len(result_df[result_df["æ¯”å¯¹çŠ¶æ€"] == COMPARE_STATUS["system_only"]]) if "æ¯”å¯¹çŠ¶æ€" in result_df.columns else 0],
            [],
            ["ã€é…ç½®ä¿¡æ¯ã€‘", ""],
            ["ä¸»é”®å­—æ®µ", config_info.get("key_columns", "")],
            ["æ•°å€¼å­—æ®µ", config_info.get("value_columns", "")],
            ["é€è§†åˆ—", config_info.get("pivot_column", "æœªä½¿ç”¨")],
            ["å·®å€¼å…¬å¼", config_info.get("diff_formula", "æ‰‹å·¥æ•°é‡ - ç³»ç»Ÿæ€»è®¡")],
            [],
            ["ã€é¢œè‰²è¯´æ˜ã€‘", ""],
            ["ç»¿è‰²", "âœ“ ä¸€è‡´"],
            ["æµ…é»„ç»¿", "â†‘ æ‰‹å·¥å¤šï¼ˆå·®å€¼>0ï¼‰"],
            ["æµ…è“", "â†“ æ‰‹å·¥å°‘ï¼ˆå·®å€¼<0ï¼‰"],
            ["æµ…çº¢", "âœ— ç¼ºå¤±ï¼ˆå•è¾¹å­˜åœ¨ï¼‰"],
        ]
        
        if pivot_values:
            data.append([])
            data.append(["ã€é€è§†å€¼ã€‘", ", ".join(pivot_values)])
        
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
